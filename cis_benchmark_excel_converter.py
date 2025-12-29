#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""CIS Benchmark to Excel Converter - Modified by kernyx64"""

import argparse
import re
import logging
import json
from pathlib import Path
from typing import Tuple, List, Dict
from collections import defaultdict

import pdfplumber
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)
logging.basicConfig(level=logging.INFO, format='%(levelname)s - %(message)s')

TITLE_PATTERN = re.compile(r'^(\d+\.\d+(?:\.\d+)*)\s*(\(L\d+\))?\s*(.*)')
PAGE_PATTERN = re.compile(r'\bPage\s+\d+\b', re.IGNORECASE)

SECTIONS = [
    'Profile Applicability:',
    'Description:',
    'Rationale:',
    'Impact:',
    'Audit:',
    'Remediation:',
    'Default Value:',
    'References:',
    'Additional Information:'
]

def clean_text(text: str) -> str:
    return PAGE_PATTERN.sub('', text)

def extract_category_names(pdf_file: Path, os_type: str = None) -> Dict[str, str]:
    """Load category names from JSON config, auto-detect OS from filename"""
    config_file = Path(__file__).parent / 'cis_categories.json'
    
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        logging.warning(f"Config file not found: {config_file}, using default")
        return config.get('default', {})
    
    # Auto-detect OS type
    if not os_type:
        filename = pdf_file.stem.lower()
        if 'debian' in filename:
            os_type = 'debian'
        elif 'ubuntu' in filename:
            os_type = 'ubuntu'
        elif 'windows_server' in filename or 'windows server' in filename:
            os_type = 'windows_server'
        elif 'windows' in filename:
            os_type = 'windows'
        else:
            os_type = 'default'
    
    categories = config.get(os_type, config.get('default', {}))
    logging.info(f"Using category mapping for: {os_type}")
    
    return categories

def get_category(number: str, categories: Dict[str, str]) -> str:
    first = number.split('.')[0]
    return categories.get(first, 'ADDITIONAL')

def extract_title_version(pdf_file: Path) -> Tuple[str, str]:
    try:
        with pdfplumber.open(str(pdf_file)) as pdf:
            lines = pdf.pages[0].extract_text().splitlines()
    except Exception as e:
        logging.error(f"Error: {e}")
        raise

    title_lines, version = [], ""
    for line in lines:
        if line.lower().startswith("v") and "-" in line:
            version = line.strip()
            break
        title_lines.append(line.strip())

    return " ".join(title_lines) or "CIS Benchmark", version

def read_pdf(pdf_file: Path, start_page: int = 10) -> str:
    logging.info(f"Reading PDF from page {start_page}...")
    try:
        with pdfplumber.open(str(pdf_file)) as pdf:
            if start_page > len(pdf.pages):
                raise ValueError(f"Start page {start_page} > total {len(pdf.pages)}")

            texts = []
            for page in tqdm(pdf.pages[start_page - 1:], desc="Extracting", unit="page"):
                texts.append(page.extract_text() or "")

    except Exception as e:
        logging.error(f"Failed: {e}")
        raise

    return "\n".join(texts)

def has_profile_applicability(lines: List[str], idx: int) -> bool:
    for i in range(idx + 1, min(idx + 10, len(lines))):
        if 'Profile Applicability:' in lines[i]:
            return True
    return False

def is_real_test(line: str) -> bool:
    """Check if line is a real test (not just a section header)"""
    # Must have (Automated) or (Manual) tag
    if '(Automated)' not in line and '(Manual)' not in line:
        return False
    
    # Extract number
    match = re.match(r'^(\d+(?:\.\d+)*)', line.strip())
    if not match:
        return False
    
    number = match.group(1)
    parts = number.split('.')
    
    # Real tests have at least 2 levels (X.Y)
    # Windows: 1.2.2 = real test
    # Debian: 1.1.1.1 = real test
    # Filter: 4.3 without tags = metadata
    if len(parts) < 2:
        return False
    
    return True

def extract_section_content(lines: List[str], idx: int) -> Tuple[str, int]:
    content = []
    i = idx + 1

    while i < len(lines):
        line = clean_text(lines[i]).strip()
        if any(line.startswith(s) for s in SECTIONS) or TITLE_PATTERN.match(line):
            break
        if line:
            content.append(line)
        i += 1

    return " ".join(content), i

def extract_recommendations(text: str) -> List[Dict]:
    logging.info("Extracting recommendations...")
    lines = text.splitlines()
    recs = []
    current = None
    i = 0

    while i < len(lines):
        line = clean_text(lines[i]).strip()

        match = TITLE_PATTERN.match(line)
        if match:
            if current:
                recs.append(current)
                current = None

            # Create temp recommendation
            if has_profile_applicability(lines, i):
                temp_rec = {
                    'Number': match.group(1),
                    'Level': match.group(2) or '',
                    'Title': match.group(3),
                }
                
                # Handle multi-line titles
                while (i + 1 < len(lines)
                       and not any(lines[i + 1].strip().startswith(s) for s in SECTIONS)
                       and not TITLE_PATTERN.match(lines[i + 1].strip())):
                    i += 1
                    temp_rec['Title'] += " " + lines[i].strip()
                
                # Check if real test after assembling full title
                full_title_line = f"{temp_rec['Number']} {temp_rec['Level']} {temp_rec['Title']}"
                if is_real_test(full_title_line):
                    current = temp_rec

        # Extract sections
        if current:
            for sec in SECTIONS:
                if line.startswith(sec):
                    content, next_i = extract_section_content(lines, i)
                    current[sec[:-1]] = content
                    i = next_i - 1
                    break

        i += 1

    if current:
        recs.append(current)

    # Deduplicate
    unique = {}
    for r in recs:
        key = (r.get('Number', ''), r.get('Title', ''))
        if key not in unique:
            unique[key] = r

    return list(unique.values())

def create_sheet(wb: Workbook, cat_name: str, recs: List[Dict], title: str, version: str, idx: int):
    name = f"{idx}. {cat_name}"
    if len(name) > 31:
        name = f"{idx}. {cat_name[:28]}"

    sheet = wb.create_sheet(title=name)

    # Title
    sheet.merge_cells('B1:H1')
    sheet['B1'].value = f"{title} - {cat_name}"
    sheet['B1'].font = Font(size=14, bold=True)
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('B2:H2')
    sheet['B2'].value = version
    sheet['B2'].font = Font(size=12, italic=True)
    sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['Number', 'Title', 'Status', 'Comments', 'Audit script', 'Remediation script',
               'Profile Applicability', 'Description', 'Rationale', 'Impact', 'Audit',
               'Remediation', 'Default Value', 'References', 'Additional Information']

    for col, h in enumerate(headers, 1):
        c = sheet.cell(row=4, column=col)
        c.value = h
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Data
    for row, rec in enumerate(recs, start=5):
        sheet.cell(row, 1, rec.get('Number', ''))
        sheet.cell(row, 2, rec.get('Title', ''))
        sheet.cell(row, 3, 'To Review')
        for col in range(4, 16):
            sheet.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')

        sheet.cell(row, 7, rec.get('Profile Applicability', ''))
        sheet.cell(row, 8, rec.get('Description', ''))
        sheet.cell(row, 9, rec.get('Rationale', ''))
        sheet.cell(row, 10, rec.get('Impact', ''))
        sheet.cell(row, 11, rec.get('Audit', ''))
        sheet.cell(row, 12, rec.get('Remediation', ''))
        sheet.cell(row, 13, rec.get('Default Value', ''))
        sheet.cell(row, 14, rec.get('References', ''))
        sheet.cell(row, 15, rec.get('Additional Information', ''))

    # Dropdown for Status column
    dv = DataValidation(type="list", formula1='"Compliant,Non-Compliant,To Review"', showDropDown=False)
    sheet.add_data_validation(dv)
    for row in range(5, len(recs) + 5):
        dv.add(sheet.cell(row, 3))

    # Conditional formatting
    start, end = 5, len(recs) + 4
    sheet.conditional_formatting.add(f"C{start}:C{end}",
        FormulaRule(formula=[f'$C{start}="Compliant"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    sheet.conditional_formatting.add(f"C{start}:C{end}",
        FormulaRule(formula=[f'$C{start}="Non-Compliant"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    sheet.conditional_formatting.add(f"C{start}:C{end}",
        FormulaRule(formula=[f'$C{start}="To Review"'],
                   fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")))

    # Column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    for col in range(7, 16):
        sheet.column_dimensions[get_column_letter(col)].width = 30
    
    # Fix row heights to 30px
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 30

def create_score_sheet(wb: Workbook, cats: List[Tuple[int, str]], num: int):
    sheet = wb.create_sheet(title=f"{num}. SCORE")

    headers = ['Section', 'Compliant', 'Non-Compliant', 'To Review', 'Total',
               '% Compliant', '% Non-Compliant', '% To Review']

    for col, h in enumerate(headers, 1):
        c = sheet.cell(1, col)
        c.value = h
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')

    for row, (idx, name) in enumerate(cats, start=2):
        ref_name = f"{idx}. {name}"
        if len(ref_name) > 31:
            ref_name = f"{idx}. {name[:28]}"
        ref = f"'{ref_name}'"

        sheet.cell(row, 1, f"{idx}. {name}")
        sheet.cell(row, 2, f'=COUNTIF({ref}!C:C, "Compliant")')
        sheet.cell(row, 3, f'=COUNTIF({ref}!C:C, "Non-Compliant")')
        sheet.cell(row, 4, f'=COUNTIF({ref}!C:C, "To Review")')
        sheet.cell(row, 5, f"=B{row}+C{row}+D{row}")
        sheet.cell(row, 6, f'=IF(E{row}=0,0,B{row}/E{row})')
        sheet.cell(row, 7, f'=IF(E{row}=0,0,C{row}/E{row})')
        sheet.cell(row, 8, f'=IF(E{row}=0,0,D{row}/E{row})')

        for col in [6, 7, 8]:
            sheet.cell(row, col).number_format = '0.00%'

    sheet.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        sheet.column_dimensions[col].width = 15

def write_excel(recs: List[Dict], output: Path, title: str, version: str, categories: Dict[str, str]):
    logging.info(f"Writing {len(recs)} recommendations...")

    # Group by category
    cats = defaultdict(list)
    for rec in recs:
        cat = get_category(rec['Number'], categories)
        cats[cat].append(rec)

    # Sort by section number
    def get_section_number(cat_name):
        if not cats[cat_name]:
            return 999
        first_rec = cats[cat_name][0]
        first_num = first_rec.get('Number', '999')
        try:
            return int(first_num.split('.')[0])
        except (ValueError, IndexError):
            return 999
    
    sorted_cats = sorted(cats.keys(), key=get_section_number)
    sorted_cats = [cat for cat in sorted_cats if cats[cat]]

    # Create Excel workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    cat_info = []
    for i, cat in enumerate(sorted_cats, 1):
        create_sheet(wb, cat, cats[cat], title, version, i)
        cat_info.append((i, cat))

    create_score_sheet(wb, cat_info, len(sorted_cats) + 1)

    wb.save(str(output))
    logging.info(f"✓ Done: {output}")

def main():
    parser = argparse.ArgumentParser(description="CIS Benchmark to Excel converter")
    parser.add_argument("-i", "--input", required=True, type=Path, help="Input PDF file")
    parser.add_argument("-o", "--output", type=Path, help="Output Excel file (default: input_name.xlsx)")
    parser.add_argument("--start_page", type=int, default=10, help="Start page for extraction (default: 10)")
    parser.add_argument("--os-type", type=str, choices=['debian', 'ubuntu', 'windows', 'windows_server', 'default'], 
                        help="OS type for category mapping (auto-detected if not specified)")
    args = parser.parse_args()

    input_file = args.input
    output_file = args.output or Path(f"{input_file.stem}.xlsx")

    title, version = extract_title_version(input_file)
    categories = extract_category_names(input_file, args.os_type)
    logging.info(f"Categories found: {categories}")
    text = read_pdf(input_file, start_page=args.start_page)
    recs = extract_recommendations(text)
    write_excel(recs, output_file, title, version, categories)

if __name__ == "__main__":
    main()
